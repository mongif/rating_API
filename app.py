from flask import Flask
from jinja2 import Template
import jinja2
import os
from flask import Flask, request, redirect, url_for
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
import string
import json
from flask import abort
from datetime import datetime

UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = set(['xlsx'])
SECRETKEY = 'qwerty'

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

def avr(d):
	summ = 0
	u = 0
	for i in d :
		try:
			summ += i["Ср.рейтинг"]
			u += 1
		except:
			pass
	return summ/u if u != 0 else 0

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

template_dir = os.path.join(os.path.dirname(__file__), 'templates' )
jinja_env = jinja2.Environment(loader = jinja2.FileSystemLoader(template_dir), autoescape = True)

def render_str(template,**params):
   t = jinja_env.get_template(template)
   return t.render(params)

def render(template, **kw):
   return render_str(template, **kw)


def file_parse():
	wb = load_workbook(filename = 'uploads\\rating_test.xlsx',data_only=True)
	sheet = wb.active
	col_names = list(string.ascii_uppercase)
	col_names.append('AA') 
	col_names.append('AB')
	col_names.append('AC')
	l = []
	d = {}
	keys = []
	for i in col_names:
	    keys.append(sheet['%s1'%i].value)
	i = 0
	for row in sheet.iter_rows('A{}:AC{}'.format(sheet.min_row + 1,sheet.max_row)):
	    i = 0
	    for cell in row:
	        d[keys[i]] = cell.value
	        i+=1
	    l.append(d)
	    d = {}
	z = avr(l)
	for i in range(len(l)):
		try:
			l[i]["Откл. от ср."]=-z+float(l[i]["Ср.рейтинг"])
		except:
			pass
	return l

@app.route("/" , methods=['GET', 'POST'])
def hello():
	if request.method == 'POST':
		if 'file' not in request.files:
			flash('No file part')
			return redirect(request.url)
		file = request.files['file']
		if file.filename == '':
			flash('No selected file')
			return redirect(request.url)
		if file and allowed_file(file.filename):
			filename = secure_filename(datetime.now().strftime(("%Y_%m_%d.xlsx")))
			file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
			return redirect(url_for('uploaded_file',filename=filename))
	else:		
		return render("index.html")
def search_name(name):
	for item in file_parse():
		if item["Фамилия, имя, отчество"] == name:
			return item
@app.route("/rating" , methods=['GET', 'POST'])
def get_data():
	name = request.args.get('name')
	key = request.args.get('key')
	if key == SECRETKEY:
		d = json.dumps(search_name(name), ensure_ascii=False).encode('utf-8')
		return d
	else:
		abort(404)
	
@app.errorhandler(404)
def page_404(e):
	return render('404.html'), 404


@app.route("/uploaded_file:<filename>" , methods=['GET', 'POST'])
def uploaded_file(filename):
	return render('success.html')
if __name__ == '__main__':
	app.run(debug = True)