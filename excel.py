from openpyxl import load_workbook
import string
wb = load_workbook(filename = 'uploads\\rating_test.xlsx',data_only=True)
sheet = wb.active
col_names = list(string.ascii_uppercase)
col_names.append('AA') 
col_names.append('AB')
col_names.append('AC')
l = []
d = {}
keys = []
for i in col_names:
    keys.append(sheet['%s1'%i].value)
i = 0
print('A{}:AC{}'.format(sheet.min_row,sheet.max_row))
for row in sheet.iter_rows('A{}:AC{}'.format(sheet.min_row + 1,sheet.max_row)):
    i = 0
    for cell in row:
        d[keys[i]] = cell.value
        i+=1
    l.append(d)
    d = {}
print(l)